import random
from urllib.parse import quote
import time
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import re

# ------------------------- Classe para Conexão com WhatsApp -------------------------
class WhatsAppBot:
    def __init__(self):
        self.navegador = None

    def iniciar(self):
        """Inicia o WhatsApp Web no navegador"""
        self.navegador = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        self.navegador.get("https://web.whatsapp.com")
        messagebox.showinfo("Escaneie o QR Code", "Escaneie o QR Code no WhatsApp Web e clique em OK.")

    def numero_valido(self):
        """Verifica se o número é válido olhando o botão de erro"""

        while len(self.navegador.find_elements(By.ID, "side")) < 1:
            time.sleep(2)

        time.sleep(random.randrange(10, 15))

        try:
            self.navegador.find_element(By.XPATH,
                                        '//*[@id="app"]/div/span[2]/div/span/div/div/div/div/div/div[2]/div/button')
            return False
        except:
            return True

    def enviar_mensagem(self, telefone, mensagem):
        self.navegador.get(f"https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}")

        if not self.numero_valido():
            raise Exception("Número inválido")

        input_box = self.navegador.find_element(By.XPATH,
                                                '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[1]/div/div/p/span')
        input_box.send_keys(Keys.ENTER)
        time.sleep(random.randrange(10, 20))

    def enviar_imagem(self, telefone, imagem_path, legenda=""):
        self.navegador.get(f"https://web.whatsapp.com/send?phone={telefone}&text={quote(legenda)}")

        if not self.numero_valido():
            raise Exception("Número inválido")

        # Continuação do envio normal
        self.navegador.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div/div[1]/button').click()

        while len(self.navegador.find_elements(By.XPATH,
                                               '//*[@id="app"]/div/span[6]/div/ul/div/div/div[2]/li/div/input')) < 1:
            time.sleep(1)

        self.navegador.find_element(By.XPATH,
                                    '//*[@id="app"]/div/span[6]/div/ul/div/div/div[2]/li/div/input').send_keys(
            imagem_path)

        while len(self.navegador.find_elements(By.XPATH,
                                               '//*[@id="app"]/div/div[3]/div/div[2]/div[2]/span/div/div/div/div[2]/div/div[2]/div[2]/div/div')) < 1:
            time.sleep(1)

        self.navegador.find_element(By.XPATH,
                                    '//*[@id="app"]/div/div[3]/div/div[2]/div[2]/span/div/div/div/div[2]/div/div[2]/div[2]/div/div').click()

        time.sleep(random.randrange(10, 20))

    def fechar(self):
        """Fecha o navegador"""
        if self.navegador:
            self.navegador.quit()

# ------------------------- Classe para Construção de Mensagens -------------------------
class MessageBuilder:
    @staticmethod
    def substituir_variaveis(texto, linha, headers):
        """Substitui variáveis {variavel} pelos valores da planilha"""
        variaveis = re.findall(r"\{(.*?)\}", texto)
        for var in variaveis:
            if var in headers:
                texto = texto.replace(f"{{{var}}}", str(linha[headers[var]].value))
            else:
                texto = texto.replace(f"{{{var}}}", f"[{var} NÃO ENCONTRADO]")
        return texto

# ------------------------- Classe para Interface Gráfica -------------------------
class AppUI:
    def __init__(self, root, main_app):
        self.root = root
        self.main_app = main_app
        self.message_list = []
        self.file_path = None

        self.criar_interface()

    def criar_interface(self):
        """Cria a interface gráfica"""
        self.root.title("WhatsApp Sender Messager")
        self.root.geometry("600x500")

        # Botão para selecionar arquivo Excel
        tk.Button(self.root, text="Selecionar Lista de Contatos", command=self.selecionar_arquivo).pack(pady=5)
        self.file_label = tk.Label(self.root, text="Nenhum arquivo selecionado", fg="gray")
        self.file_label.pack()

        # Área para adicionar mensagens
        tk.Label(self.root, text="Digite uma mensagem (use {variavel} para personalizar):", font=("Arial", 10)).pack(pady=5)
        self.text_entry = tk.Text(self.root, height=3, width=50)
        self.text_entry.pack()

        tk.Button(self.root, text="➕ Adicionar Texto", command=self.adicionar_texto, bg="lightblue").pack(pady=5)
        tk.Button(self.root, text="🖼️ Adicionar Imagem", command=self.adicionar_imagem, bg="lightgreen").pack(pady=5)

        # Área de visualização da sequência
        tk.Label(self.root, text="📜 Sequência das Mensagens:", font=("Arial", 10, "bold")).pack(pady=5)
        self.preview_frame = tk.Frame(self.root)
        self.preview_frame.pack(fill="both", expand=True)

        # Botão para enviar mensagens
        tk.Button(self.root, text="🚀 Enviar Mensagens", command=self.main_app.enviar_mensagens, bg="green",
                  fg="white").pack(pady=20)

    def selecionar_arquivo(self):
        """Seleciona o arquivo de contatos"""
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        self.file_label.config(text=f"Arquivo: {self.file_path.split('/')[-1]}" if self.file_path else "Nenhum arquivo selecionado")

    def adicionar_texto(self):
        """Adiciona um texto à lista de mensagens e atualiza a interface"""
        text = self.text_entry.get("1.0", tk.END).strip()
        if text:
            self.message_list.append({"type": "text", "content": text})
            self.text_entry.delete("1.0", tk.END)
            self.atualizar_lista_mensagens()

    def adicionar_imagem(self):
        """Adiciona uma imagem e permite adicionar uma legenda"""
        file_path = filedialog.askopenfilename(filetypes=[("Imagens", "*.jpg;*.jpeg;*.png")])
        if file_path:
            legenda = simpledialog.askstring("Legenda", "Digite a legenda para a imagem (opcional):")
            self.message_list.append({"type": "image", "content": file_path, "caption": legenda if legenda else ""})
            self.atualizar_lista_mensagens()

    def atualizar_lista_mensagens(self):
        """Atualiza a visualização da sequência de mensagens"""
        for widget in self.preview_frame.winfo_children():
            widget.destroy()

        for i, msg in enumerate(self.message_list):
            frame = tk.Frame(self.preview_frame, relief=tk.RAISED, borderwidth=1)
            frame.pack(fill="x", pady=2)

            label = tk.Label(frame,
                             text=f"{'📜 Texto' if msg['type'] == 'text' else '🖼️ Imagem'}: {msg['content'][:30]}...")
            label.pack(side="left", padx=5)

            tk.Button(frame, text="⬆", command=lambda idx=i: self.move_up(idx)).pack(side="right")
            tk.Button(frame, text="⬇", command=lambda idx=i: self.move_down(idx)).pack(side="right")
            tk.Button(frame, text="❌", command=lambda idx=i: self.remove_message(idx)).pack(side="right")

    def move_up(self, index):
        """Move um item para cima"""
        if index > 0:
            self.message_list[index], self.message_list[index - 1] = self.message_list[index - 1], self.message_list[index]
            self.atualizar_lista_mensagens()

    def move_down(self, index):
        """Move um item para baixo"""
        if index < len(self.message_list) - 1:
            self.message_list[index], self.message_list[index + 1] = self.message_list[index + 1], self.message_list[index]
            self.atualizar_lista_mensagens()

    def remove_message(self, index):
        """Remove um item da lista"""
        del self.message_list[index]
        self.atualizar_lista_mensagens()


# ------------------------- Classe Principal -------------------------
class MainApp:
    def __init__(self):
        self.whatsapp_bot = WhatsAppBot()
        self.root = tk.Tk()
        self.ui = AppUI(self.root, self)

    def run(self):
        """Inicia a interface gráfica"""
        self.root.mainloop()

    def enviar_mensagens(self):
        """Lê a planilha e envia as mensagens para cada contato"""
        if not self.ui.file_path:
            messagebox.showerror("Erro", "Selecione um arquivo de contatos!")
            return

        if not self.ui.message_list:
            messagebox.showerror("Erro", "Adicione pelo menos uma mensagem ou imagem!")
            return

        self.whatsapp_bot.iniciar()

        contatos_df = openpyxl.load_workbook(self.ui.file_path)
        pagina_contatos = contatos_df.active

        headers = {cell.value: index for index, cell in enumerate(pagina_contatos[1])}

        # Criar um arquivo para erros, se necessário
        erro_wb = openpyxl.Workbook()
        erro_ws = erro_wb.active
        erro_ws.append([cell.value for cell in pagina_contatos[1]])  # Cabeçalho

        for linha in pagina_contatos.iter_rows(min_row=2):
            telefone = linha[headers.get("Telefone", 1)].value
            try:
                for msg in self.ui.message_list:
                    if msg["type"] == "text":
                        mensagem = MessageBuilder.substituir_variaveis(msg["content"], linha, headers)
                        self.whatsapp_bot.enviar_mensagem(telefone, mensagem)

                    elif msg["type"] == "image":
                        legenda = MessageBuilder.substituir_variaveis(msg["caption"], linha, headers) if msg[
                            "caption"] else ""
                        self.whatsapp_bot.enviar_imagem(telefone, msg["content"], legenda)

            except Exception as e:
                print(f"Erro com o número {telefone}: {e}")
                erro_ws.append([cell.value for cell in linha])

        self.whatsapp_bot.fechar()
        # Salvar erros se houver falhas
        erro_filename = "erros_envio.xlsx"
        if len(erro_ws["A"]) > 1:  # Se houver erros registrados
            erro_wb.save(erro_filename)
            messagebox.showwarning("Envio Concluído",
                                f"Mensagens enviadas, mas alguns contatos falharam. Veja {erro_filename}.")
        else:
            messagebox.showinfo("Concluído", "Mensagens enviadas com sucesso!")

    def run(self):
        """Inicia a interface gráfica"""
        self.root.mainloop()

# ------------------------- Iniciar a Aplicação -------------------------
if __name__ == "__main__":
    app = MainApp()
    app.run()